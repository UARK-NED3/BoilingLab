"""Compare boiling curves across multiple pool boiling tests.

The default comparison uses Boiling-412, Boiling-413, Boiling-416, and
Boiling-417 as a subcooled flat-copper heat-load sweep.
"""

from __future__ import annotations

import argparse
import csv
import json
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from pyXSteam.XSteam import XSteam

from run_single_case_demo import compute_temperature_quantities, parse_lvm_start_time_seconds, read_lvm


DEFAULT_CASES = ["Boiling-412", "Boiling-413", "Boiling-416", "Boiling-417"]


def normalize_test_id(test_id: str) -> str:
    return test_id if test_id.startswith("Boiling-") else f"Boiling-{test_id}"


def load_test_log(path: Path) -> dict[str, dict[str, object]]:
    if not path.exists():
        return {}
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: dict[str, dict[str, object]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        record = dict(zip(headers, row))
        rows[str(row[0])] = record
    return rows


def pressure_stats(folder: Path, target_pressure_kpa: float) -> dict[str, float]:
    count = 0
    total = 0.0
    total_sq = 0.0
    error_sq = 0.0
    min_value = np.inf
    max_value = -np.inf

    for chunk in pd.read_csv(folder / "Pressure.lvm", skiprows=22, sep="\t", chunksize=1_000_000):
        pressure = chunk["Voltage"].to_numpy(dtype=float)
        count += pressure.size
        total += float(np.sum(pressure))
        total_sq += float(np.sum(pressure**2))
        error_sq += float(np.sum((pressure - target_pressure_kpa) ** 2))
        min_value = min(min_value, float(np.nanmin(pressure)))
        max_value = max(max_value, float(np.nanmax(pressure)))

    mean = total / count
    variance = (total_sq - count * mean**2) / (count - 1)
    return {
        "pressure_mean_kPa": float(
            Decimal(str(mean)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        ),
        "pressure_std_kPa": float(np.sqrt(max(variance, 0.0))),
        "pressure_rmse_vs_target_kPa": float(np.sqrt(error_sq / count)),
        "pressure_min_kPa": min_value,
        "pressure_max_kPa": max_value,
    }


def heating_mask_from_dc_power(folder: Path, time_s: np.ndarray, threshold_w: float) -> tuple[np.ndarray, dict[str, float]]:
    dc_path = folder / "DC_power.lvm"
    if not dc_path.exists():
        raise FileNotFoundError(f"DC power file not found: {dc_path}")

    dc = read_lvm(folder, "DC_power.lvm")
    dc_time = dc.iloc[:, 0].to_numpy(dtype=float)
    dc_power = dc.iloc[:, -1].to_numpy(dtype=float)

    time_offset = parse_lvm_start_time_seconds(dc_path) - parse_lvm_start_time_seconds(folder / "Temperature.lvm")
    aligned_dc_time = dc_time + time_offset

    finite_dc = np.isfinite(aligned_dc_time) & np.isfinite(dc_power)
    aligned_dc_time = aligned_dc_time[finite_dc]
    dc_power = dc_power[finite_dc]
    order = np.argsort(aligned_dc_time)
    aligned_dc_time = aligned_dc_time[order]
    dc_power = dc_power[order]

    interpolated_power = np.interp(time_s, aligned_dc_time, dc_power, left=0.0, right=0.0)
    heating_mask = interpolated_power > threshold_w
    if not np.any(heating_mask):
        raise ValueError(f"No heating samples found in {folder} using power threshold {threshold_w} W")

    heating_times = time_s[heating_mask]
    return heating_mask, {
        "heating_start_time_s": float(np.nanmin(heating_times)),
        "heating_end_time_s": float(np.nanmax(heating_times)),
        "heating_duration_s": float(np.nanmax(heating_times) - np.nanmin(heating_times)),
        "max_dc_power_W": float(np.nanmax(dc_power)),
        "mean_dc_power_during_heating_W": float(np.nanmean(interpolated_power[heating_mask])),
    }


def load_curve_case(
    test_id: str,
    raw_root: Path,
    test_log: dict[str, dict[str, object]],
    target_pressure_kpa: float,
    power_threshold_w: float,
) -> dict[str, object]:
    test_id = normalize_test_id(test_id)
    folder = raw_root / test_id
    if not folder.exists():
        raise FileNotFoundError(f"Raw folder not found: {folder}")

    temp = read_lvm(folder, "Temperature.lvm").rename(columns={"X_Value": "Time (sec)"})
    temp_data = compute_temperature_quantities(temp)
    stats = pressure_stats(folder, target_pressure_kpa)
    t_sat_c = XSteam(XSteam.UNIT_SYSTEM_BARE).tsat_p(stats["pressure_mean_kPa"] / 1000.0) - 273.15

    heat_flux = temp_data["heat_flux"]
    surface_temperature = temp_data["surface_temperature"]
    wall_superheat = surface_temperature - t_sat_c
    time_s = temp_data["time_s"]
    heating_mask, heating_stats = heating_mask_from_dc_power(folder, time_s, power_threshold_w)

    log_row = test_log.get(test_id, {})
    power_value = log_row.get("MagnaDC \nPower (W)")
    try:
        power_w = float(power_value) if power_value is not None else None
    except (TypeError, ValueError):
        power_w = None

    return {
        "test_id": test_id,
        "folder": str(folder),
        "time_s": time_s[heating_mask],
        "surface_temperature_C": surface_temperature[heating_mask],
        "wall_superheat_C": wall_superheat[heating_mask],
        "heat_flux_W_cm2": heat_flux[heating_mask],
        "T_sat_C": float(t_sat_c),
        "mean_liquid_temp_C": float(np.nanmean(temp["Liquid Temp"].to_numpy(dtype=float))),
        "mean_vapour_temp_C": float(np.nanmean(temp["Vapour Temp"].to_numpy(dtype=float))),
        "max_heat_flux_W_cm2": float(np.nanmax(heat_flux[heating_mask])),
        "time_at_max_heat_flux_s": float(time_s[heating_mask][np.nanargmax(heat_flux[heating_mask])]),
        "max_surface_temp_C": float(np.nanmax(surface_temperature[heating_mask])),
        "time_at_max_surface_temp_s": float(
            time_s[heating_mask][np.nanargmax(surface_temperature[heating_mask])]
        ),
        "applied_power_W": power_w,
        "status": log_row.get("Status"),
        "surface": log_row.get("Surface"),
        **heating_stats,
        **stats,
    }


def case_label(case: dict[str, object]) -> str:
    power = case.get("applied_power_W")
    if power is None:
        return str(case["test_id"])
    return f"$P_{{\\mathrm{{load}}}}$ = {power:g} W"


def plot_comparison(
    cases: list[dict[str, object]],
    x_key: str,
    x_label: str,
    output_path: Path,
    downsample: int,
) -> None:
    fig, ax = plt.subplots(figsize=(9, 8))
    for case in cases:
        x = np.asarray(case[x_key], dtype=float)
        y = np.asarray(case["heat_flux_W_cm2"], dtype=float)
        finite = np.isfinite(x) & np.isfinite(y)
        step = max(1, downsample)
        line = ax.plot(x[finite][::step], y[finite][::step], linewidth=2.4, label=case_label(case))[0]
        power = case.get("applied_power_W")
        if power is not None:
            ax.axhline(
                y=float(power),
                color=line.get_color(),
                linestyle="--",
                linewidth=1.8,
                alpha=0.75,
            )

    ax.set_xlabel(x_label, fontsize=22)
    ax.set_ylabel("Heat flux, $q^{\\prime\\prime}$ (W/cm²)", fontsize=22)
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.tick_params(axis="both", which="major", labelsize=18)
    ax.legend(fontsize=18)
    fig.tight_layout()
    fig.savefig(output_path, dpi=220)
    plt.close(fig)


def write_summary(output_dir: Path, cases: list[dict[str, object]]) -> None:
    scalar_keys = [
        "test_id",
        "applied_power_W",
        "status",
        "surface",
        "pressure_mean_kPa",
        "pressure_std_kPa",
        "T_sat_C",
        "mean_liquid_temp_C",
        "mean_vapour_temp_C",
        "max_heat_flux_W_cm2",
        "time_at_max_heat_flux_s",
        "max_surface_temp_C",
        "time_at_max_surface_temp_s",
        "heating_start_time_s",
        "heating_end_time_s",
        "heating_duration_s",
        "mean_dc_power_during_heating_W",
    ]
    rows = [{key: case.get(key) for key in scalar_keys} for case in cases]
    with (output_dir / "summary.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=scalar_keys)
        writer.writeheader()
        writer.writerows(rows)

    (output_dir / "summary.json").write_text(json.dumps(rows, indent=2), encoding="utf-8")

    markdown_rows = "\n".join(
        "| "
        + " | ".join(
            [
                str(row["test_id"]),
                "" if row["applied_power_W"] is None else f"{row['applied_power_W']:g}",
                str(row["status"]),
                f"{row['pressure_mean_kPa']:.2f}",
                f"{row['mean_liquid_temp_C']:.2f}",
                f"{row['max_heat_flux_W_cm2']:.2f}",
                f"{row['max_surface_temp_C']:.2f}",
            ]
        )
        + " |"
        for row in rows
    )
    (output_dir / "README.md").write_text(
        "# Multi-Case Boiling Curve Comparison\n\n"
        "This generated comparison plots heat flux against wall temperature and "
        "wall superheat for the selected tests. For these demo cases, the test "
        "log marks CHF as not reached, so the plots are intended for curve-shape "
        "comparison rather than CHF ranking.\n\n"
        "| Test ID | Power (W) | Status | Mean pressure (kPa) | Mean liquid temp (C) | "
        "Max heat flux (W/cm^2) | Max surface temp (C) |\n"
        "| --- | ---: | --- | ---: | ---: | ---: | ---: |\n"
        f"{markdown_rows}\n\n"
        "Only samples with positive DC output power are included in these generated curves.\n\n"
        "Generated plots:\n\n"
        "- [Heat flux vs wall temperature](plots/heat_flux_vs_wall_temperature.png)\n"
        "- [Heat flux vs wall superheat](plots/heat_flux_vs_wall_superheat.png)\n",
        encoding="utf-8",
    )


def write_curve_csv(output_dir: Path, cases: list[dict[str, object]], downsample: int) -> None:
    with (output_dir / "curves.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "test_id",
                "time_s",
                "surface_temperature_C",
                "wall_superheat_C",
                "heat_flux_W_cm2",
            ]
        )
        step = max(1, downsample)
        for case in cases:
            for values in zip(
                case["time_s"][::step],
                case["surface_temperature_C"][::step],
                case["wall_superheat_C"][::step],
                case["heat_flux_W_cm2"][::step],
            ):
                writer.writerow([case["test_id"], *values])


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--test-ids", nargs="+", default=DEFAULT_CASES)
    parser.add_argument("--raw-root", default=r"Y:\0_Ishraq\New Pool Boiling Video")
    parser.add_argument("--test-log", default=r"metadata\Pool Boiling Test Log.xlsx")
    parser.add_argument("--output-dir", default=r"demos\Boiling-412-413-416-417\generated")
    parser.add_argument("--target-pressure", type=float, default=97.7)
    parser.add_argument(
        "--power-threshold-w",
        type=float,
        default=0.0,
        help="Keep only temperature samples whose aligned DC output power is greater than this threshold.",
    )
    parser.add_argument(
        "--downsample",
        type=int,
        default=5,
        help="Plot/export every Nth temperature sample. Use 1 for all points.",
    )
    return parser


def main() -> None:
    args = build_parser().parse_args()
    output_dir = Path(args.output_dir)
    plots_dir = output_dir / "plots"
    plots_dir.mkdir(parents=True, exist_ok=True)

    test_log = load_test_log(Path(args.test_log))
    cases = [
        load_curve_case(test_id, Path(args.raw_root), test_log, args.target_pressure, args.power_threshold_w)
        for test_id in args.test_ids
    ]

    plot_comparison(
        cases,
        "surface_temperature_C",
        "Wall temperature, $T_{\\mathrm{w}}$ (°C)",
        plots_dir / "heat_flux_vs_wall_temperature.png",
        args.downsample,
    )
    plot_comparison(
        cases,
        "wall_superheat_C",
        "Wall superheat, $\\Delta T$ (°C)",
        plots_dir / "heat_flux_vs_wall_superheat.png",
        args.downsample,
    )
    write_summary(output_dir, cases)
    write_curve_csv(output_dir, cases, args.downsample)
    print(json.dumps([{k: v for k, v in case.items() if not isinstance(v, np.ndarray)} for case in cases], indent=2))


if __name__ == "__main__":
    main()
